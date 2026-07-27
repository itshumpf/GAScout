"""Build counties/gwinnett.json from the Gwinnett Tax Commissioner weekly
delinquent list (.xlsx). AGGREGATE ONLY — no owner names or addresses are
ever written out (same PII rule as DeKalb). Re-run when the weekly file
updates.

Usage: python build_gwinnett.py "/path/to/WeeklyDelqLender-*.xlsx"
"""
import sys, json, re
from collections import defaultdict
import openpyxl

def num(x):
    try: return float(x)
    except (TypeError, ValueError): return 0.0

def build(xlsx_path, out_path, run_date="2026-07-21"):
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    total_rows = 0; total_bal = 0.0
    parcels = set(); owners = set()
    by_year = defaultdict(lambda: [0, 0.0])
    real_rows = 0; real_bal = 0.0; real_parcels = set()
    bands = {"$0–100": 0, "$100–500": 0, "$500–1K": 0, "$1K–5K": 0, "$5K+": 0}
    bals = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        parcel, owner, situs, year, *_rest = r
        totbal = r[9]
        if parcel is None:
            continue
        b = num(totbal)
        total_rows += 1; total_bal += b; bals.append(b)
        parcels.add(parcel); owners.add(owner)
        y = str(year).split(".")[0] if year is not None else "?"
        by_year[y][0] += 1; by_year[y][1] += round(b, 2)
        if str(parcel)[0].upper() == "R":
            real_rows += 1; real_bal += b; real_parcels.add(parcel)
        if   b < 100:  bands["$0–100"] += 1
        elif b < 500:  bands["$100–500"] += 1
        elif b < 1000: bands["$500–1K"] += 1
        elif b < 5000: bands["$1K–5K"] += 1
        else:          bands["$5K+"] += 1
    bals.sort()
    median = round(bals[len(bals)//2], 2) if bals else 0.0
    largest = round(max(bals), 2) if bals else 0.0
    by_year_out = {y: {"records": c, "due": round(v, 2)}
                   for y, (c, v) in sorted(by_year.items()) if y.isdigit()}

    data = {
        "county": "gwinnett", "name": "Gwinnett", "fips": "13135",
        "seat": "Lawrenceville", "is_live": True,
        "status": "LIVE PRODUCTION", "has_amounts": True,
        "adapter": "gwinnett_xlsx",
        "legal_organ": "Gwinnett County Tax Commissioner",
        "source": "Gwinnett County Tax Commissioner — Weekly Delinquent Tax List (.xlsx)",
        "run_date": run_date, "generated_at": run_date,
        "records": total_rows,
        "owners": len(owners),
        "parcels": len(parcels),
        "total_due": round(total_bal, 2),
        "by_year": by_year_out,
        "amount_bands": bands,
        "flags": None,
        "median_due": median,
        "largest_single_bill": largest,
        "real_property": {"line_items": real_rows, "parcels": len(real_parcels),
                          "total_due": round(real_bal, 2)},
        "note": ("Full delinquent roll (real + personal property) from the county's "
                 "weekly published spreadsheet. Aggregate figures only — no owner "
                 "names or addresses are published. Real property (parcels prefixed 'R') "
                 f"is {real_rows:,} line items across {len(real_parcels):,} parcels, "
                 f"${round(real_bal,2):,.2f} of the total.")
    }
    with open(out_path, "w", encoding="utf-8") as fh:
        json.dump(data, fh, indent=2)
    return data

if __name__ == "__main__":
    xlsx = sys.argv[1]
    out = sys.argv[2] if len(sys.argv) > 2 else "src/data/counties/gwinnett.json"
    d = build(xlsx, out)
    print(f"wrote {out}")
    print(f"records={d['records']:,} owners={d['owners']:,} parcels={d['parcels']:,} "
          f"total_due=${d['total_due']:,.2f}")
    print(f"real property: {d['real_property']['parcels']:,} parcels "
          f"${d['real_property']['total_due']:,.2f}")
